from pathlib import Path
import os
import sys
import ipaddress
from typing import Optional, Literal
from datetime import datetime
from io import BytesIO

from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, status, Depends, Header, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, EmailStr, ConfigDict, field_validator
from psycopg2.extras import RealDictCursor
from supabase import create_client, Client
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Resolve paths and load environment variables
BASE_DIR = Path(__file__).resolve().parent
ROOT_DIR = BASE_DIR.parent
if str(ROOT_DIR) not in sys.path:
    sys.path.append(str(ROOT_DIR))

STATIC_DIR = BASE_DIR / "static"
load_dotenv(ROOT_DIR / ".env")

# Support running as a module (`python -m backend.main`) and as a script (`python backend/main.py`)
try:  # pragma: no cover - import resolution guard
    from .database import execute_query, get_db_connection
except ImportError:  # when executed without package context
    from database import execute_query, get_db_connection

app = FastAPI(title="Attendance Management API")

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, specify exact origins
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize Supabase clients with fallback env var names
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY") or os.getenv("SUPABASE_ANON_KEY")  # anon/public key
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY") or os.getenv("SUPABASE_SERVICE_KEY")
DEBUG_AUTH = os.getenv("DEBUG_AUTH", "false").lower() == "true"

if not SUPABASE_URL or not SUPABASE_KEY:
    raise ValueError("SUPABASE_URL and SUPABASE_KEY (or SUPABASE_ANON_KEY) must be set in environment variables")

if DEBUG_AUTH:
    print(f"[AUTH DEBUG] SUPABASE_URL set={bool(SUPABASE_URL)}")
    print(f"[AUTH DEBUG] SUPABASE_KEY present={bool(SUPABASE_KEY)} len={len(SUPABASE_KEY) if SUPABASE_KEY else 0}")
    print(f"[AUTH DEBUG] SUPABASE_SERVICE_ROLE_KEY present={bool(SUPABASE_SERVICE_ROLE_KEY)}")

# Use anon key for user-facing actions (sign-up/sign-in)
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# Optional service role client (only if key is provided)
supabase_admin: Optional[Client] = None
if SUPABASE_SERVICE_ROLE_KEY:
    supabase_admin = create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)
elif DEBUG_AUTH:
    print("[AUTH DEBUG] SUPABASE_SERVICE_ROLE_KEY missing; admin client disabled")

# Mount static files
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

# Office IP Configuration - Load from environment or use defaults
# You can set OFFICE_IP_WHITELIST in .env as comma-separated IPs/CIDRs
# Example: OFFICE_IP_WHITELIST=192.168.1.0/24,10.0.0.0/8,127.0.0.1
OFFICE_IP_WHITELIST = os.getenv("OFFICE_IP_WHITELIST", "127.0.0.1,::1,192.168.1.0/24")
ALLOWED_OFFICE_IPS = [ip.strip() for ip in OFFICE_IP_WHITELIST.split(",") if ip.strip()]

def get_client_ip(request: Request) -> str:
    """Extract client IP address from request, checking X-Forwarded-For and X-Real-IP headers"""
    # Check X-Forwarded-For header (used by proxies like nginx)
    forwarded_for = request.headers.get("X-Forwarded-For")
    if forwarded_for:
        # Get the first IP in the chain (original client IP)
        return forwarded_for.split(",")[0].strip()
    
    # Check X-Real-IP header
    real_ip = request.headers.get("X-Real-IP")
    if real_ip:
        return real_ip.strip()
    
    # Fall back to direct client IP
    if request.client:
        return request.client.host
    
    return "unknown"

def is_office_network(ip_address: str) -> bool:
    """Check if the given IP address is within the office network whitelist"""
    if not ALLOWED_OFFICE_IPS:
        # If no whitelist configured, allow all (for development)
        return True
    
    try:
        client_ip = ipaddress.ip_address(ip_address)
        
        for allowed in ALLOWED_OFFICE_IPS:
            try:
                # Check if it's a network range (CIDR notation)
                if "/" in allowed:
                    if client_ip in ipaddress.ip_network(allowed, strict=False):
                        return True
                # Check if it's a single IP
                else:
                    if client_ip == ipaddress.ip_address(allowed):
                        return True
            except ValueError:
                # Invalid IP format in whitelist, skip
                continue
        
        return False
    except ValueError:
        # Invalid client IP format
        return False

def require_office_network(request: Request):
    """Dependency to ensure request comes from office network"""
    client_ip = get_client_ip(request)
    
    if not is_office_network(client_ip):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=f"Access denied. Check-in is only allowed from office network. Your IP: {client_ip}"
        )
    
    return client_ip


@app.on_event("startup")
async def ensure_schema():
    """Ensure required tables exist in the database."""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('CREATE EXTENSION IF NOT EXISTS "pgcrypto";')

            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS profiles (
                    id UUID PRIMARY KEY,
                    email TEXT UNIQUE NOT NULL,
                    full_name TEXT,
                    role TEXT NOT NULL DEFAULT 'member' CHECK (role IN ('admin', 'member')),
                    created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                )
                """
            )

            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS attendance (
                    id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
                    user_id UUID NOT NULL REFERENCES profiles(id) ON DELETE CASCADE,
                    check_in_time TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    check_out_time TIMESTAMPTZ,
                    status TEXT NOT NULL DEFAULT 'present' CHECK (status IN ('present', 'absent', 'late')),
                    notes TEXT,
                    latitude DOUBLE PRECISION,
                    longitude DOUBLE PRECISION,
                    created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                )
                """
            )
            
            # Add latitude and longitude columns if they don't exist (for existing tables)
            cursor.execute(
                """
                DO $$ 
                BEGIN
                    IF NOT EXISTS (SELECT 1 FROM information_schema.columns 
                                   WHERE table_name='attendance' AND column_name='latitude') THEN
                        ALTER TABLE attendance ADD COLUMN latitude DOUBLE PRECISION;
                    END IF;
                    IF NOT EXISTS (SELECT 1 FROM information_schema.columns 
                                   WHERE table_name='attendance' AND column_name='longitude') THEN
                        ALTER TABLE attendance ADD COLUMN longitude DOUBLE PRECISION;
                    END IF;
                END $$;
                """
            )

            cursor.execute("CREATE INDEX IF NOT EXISTS idx_attendance_user_id ON attendance(user_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_attendance_check_in_time ON attendance(check_in_time)")

            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS leaves (
                    id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
                    user_id UUID NOT NULL REFERENCES profiles(id) ON DELETE CASCADE,
                    start_date DATE NOT NULL,
                    end_date DATE NOT NULL,
                    reason TEXT NOT NULL,
                    status TEXT NOT NULL DEFAULT 'pending' CHECK (status IN ('pending', 'approved', 'rejected')),
                    admin_note TEXT,
                    approved_by UUID REFERENCES profiles(id),
                    created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                )
                """
            )

            cursor.execute("CREATE INDEX IF NOT EXISTS idx_leaves_user_id ON leaves(user_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_leaves_status ON leaves(status)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_leaves_start_date ON leaves(start_date)")

            conn.commit()
            cursor.close()
    except Exception as e:
        if DEBUG_AUTH:
            print(f"[DB INIT] Failed to ensure schema: {e}")


# Authentication and Authorization utilities
# Updated: Fixed token validation
async def get_current_user(authorization: Optional[str] = Header(None, alias="Authorization")):
    """Dependency to get current authenticated user"""
    if not authorization:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Not authenticated",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    try:
        # Remove 'Bearer ' prefix if present
        token = authorization.replace("Bearer ", "").strip()
        if DEBUG_AUTH:
            print(f"[AUTH DEBUG] Incoming token prefix: {token[:12]}... (len={len(token)})")
        
        # Validate token using anon/public client (works with access token)
        response = supabase.auth.get_user(token)
        if DEBUG_AUTH:
            print(f"[AUTH DEBUG] get_user response: user={bool(response and response.user)}")
        
        if not response or not response.user:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid authentication credentials",
                headers={"WWW-Authenticate": "Bearer"},
            )
        return response
    except HTTPException:
        raise
    except Exception as e:
        if DEBUG_AUTH:
            print(f"[AUTH DEBUG] Exception during auth: {e}")
        msg = str(e)
        if "api key" in msg.lower():
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Server auth misconfigured: invalid Supabase API key. Check SUPABASE_URL and SUPABASE_ANON_KEY in .env."
            )
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail=f"Authentication failed: {msg}",
            headers={"WWW-Authenticate": "Bearer"},
        )


def require_admin(user = Depends(get_current_user)):
    """Dependency to check if user is an admin"""
    # Get role from database for accuracy
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            cursor.execute("SELECT role FROM profiles WHERE id = %s", (user.user.id,))
            profile = cursor.fetchone()
            cursor.close()
            
            if profile:
                user_role = profile.get("role", "member")
            else:
                user_role = user.user.user_metadata.get("role", "member")
    except Exception as e:
        print(f"⚠️ Role check error: {e}")
        user_role = user.user.user_metadata.get("role", "member")
    
    if user_role != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You don't have permission to access this resource. Admin access required."
        )
    
    return user


# Trigger reload - Update timestamp or comment
# Last updated: 2026-01-21 - Fixed role handling

# Pydantic models
class SignUpRequest(BaseModel):
    model_config = ConfigDict(
        json_schema_extra={
            "example": {
                "email": "user@example.com",
                "password": "securePassword123",
                "full_name": "John Doe",
                "role": "member"
            }
        }
    )
    
    email: EmailStr
    password: str
    full_name: Optional[str] = None
    role: Literal["member", "admin"] = "member"
    
    @field_validator('role')
    @classmethod
    def validate_role(cls, v):
        if v not in ["member", "admin"]:
            raise ValueError('Role must be either "member" or "admin"')
        return v


class SignInRequest(BaseModel):
    model_config = ConfigDict(
        json_schema_extra={
            "example": {
                "email": "user@example.com",
                "password": "securePassword123"
            }
        }
    )
    
    email: EmailStr
    password: str


class AttendanceCheckoutRequest(BaseModel):
    attendance_id: str


class CheckInRequest(BaseModel):
    latitude: Optional[float] = None
    longitude: Optional[float] = None


class LeaveRequest(BaseModel):
    start_date: str  # Format: YYYY-MM-DD
    end_date: str    # Format: YYYY-MM-DD
    reason: str


class LeaveApprovalRequest(BaseModel):
    leave_id: str
    status: Literal["approved", "rejected"]
    admin_note: Optional[str] = None


class AuthResponse(BaseModel):
    access_token: str
    token_type: str
    user: dict
    message: str


@app.get("/", include_in_schema=False)
async def serve_frontend():
    """Serve the frontend HTML"""
    return FileResponse(STATIC_DIR / "index.html")


@app.get("/api")
async def root():
    """API Root endpoint"""
    return {
        "message": "Welcome to Attendance Management API",
        "version": "1.0.0",
        "endpoints": {
            "signup": "/auth/signup",
            "signin": "/auth/signin",
            "health": "/health",
            "docs": "/docs"
        }
    }


@app.get("/health")
async def health_check():
    """Health check endpoint"""
    return {
        "status": "healthy",
        "service": "Attendance Management API",
        "supabase": "connected"
    }


@app.post("/auth/signup", response_model=AuthResponse, status_code=status.HTTP_201_CREATED)
async def sign_up(user_data: SignUpRequest):
    """
    Register a new user with email and password
    """
    try:
        # Sign up the user with Supabase
        credentials_data = {
            "email": user_data.email,
            "password": user_data.password
        }
        
        # Add user metadata including role and full_name
        metadata = {"role": user_data.role}
        if user_data.full_name:
            metadata["full_name"] = user_data.full_name
        credentials_data["data"] = metadata
        
        print(f"===== SIGNUP DEBUG =====")
        print(f"Signing up user: {user_data.email}")
        print(f"Requested role: {user_data.role}")
        print(f"Full name: {user_data.full_name}")
        print(f"Metadata to save: {metadata}")
        print(f"========================")
        
        response = supabase.auth.sign_up(credentials=credentials_data)

        if DEBUG_AUTH:
            print(f"[AUTH DEBUG] sign_up email={user_data.email} ok={bool(response and response.user)}")

        if not response.user:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Sign up failed. User may already exist or invalid credentials."
            )

        print(f"✅ User created successfully")
        print(f"User ID: {response.user.id}")
        print(f"User metadata saved: {response.user.user_metadata}")
        
        # Insert user profile into database
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(
                    """
                    INSERT INTO profiles (id, email, full_name, role)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (id) DO UPDATE SET
                        full_name = EXCLUDED.full_name,
                        role = EXCLUDED.role,
                        updated_at = NOW()
                    """,
                    (response.user.id, user_data.email, user_data.full_name, user_data.role)
                )
                conn.commit()
                cursor.close()
                print(f"✅ Profile created in database with role: {user_data.role}")
        except Exception as e:
            print(f"⚠️ Failed to create profile in database: {e}")
        
        print(f"========================")

        return AuthResponse(
            access_token=response.session.access_token if response.session else "",
            token_type="bearer",
            user={
                "id": response.user.id,
                "email": response.user.email,
                "full_name": user_data.full_name,
                "role": user_data.role,
                "user_metadata": metadata
            },
            message="User registered successfully. Please check your email to verify your account."
        )

    except HTTPException:
        raise
    except Exception as e:
        msg = str(e)
        if "already registered" in msg.lower() or "user already exists" in msg.lower():
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="User already registered. Please sign in instead."
            )
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Sign up failed: {msg}"
        )


@app.post("/auth/signin", response_model=AuthResponse)
async def sign_in(credentials: SignInRequest):
    """
    Sign in an existing user with email and password
    """
    try:
        # Sign in the user with Supabase
        response = supabase.auth.sign_in_with_password(
            credentials={
                "email": credentials.email,
                "password": credentials.password
            }
        )

        if DEBUG_AUTH:
            print(f"[AUTH DEBUG] sign_in_with_password email={credentials.email} ok={bool(response and response.user and response.session)}")

        if not response.user or not response.session:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid email or password"
            )

        # Extract role from user metadata, default to 'member' if not set
        # This handles both new users (with role) and existing users (without role)
        user_metadata = response.user.user_metadata or {}
        user_role = user_metadata.get("role", "member")
        user_full_name = user_metadata.get("full_name", "")
        
        print(f"===== SIGNIN DEBUG =====")
        print(f"User email: {response.user.email}")
        print(f"User ID: {response.user.id}")
        print(f"Raw user_metadata: {user_metadata}")
        
        # Get role from database profiles table
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT role, full_name FROM profiles WHERE id = %s",
                    (response.user.id,)
                )
                profile = cursor.fetchone()
                cursor.close()
                
                if profile:
                    user_role = profile[0]
                    user_full_name = profile[1] or user_full_name
                    print(f"✅ Role from database: {user_role}")
                else:
                    # Create profile if it doesn't exist (for existing users)
                    print(f"⚠️ No profile found, creating with default 'member' role")
                    cursor = conn.cursor()
                    cursor.execute(
                        """
                        INSERT INTO profiles (id, email, full_name, role)
                        VALUES (%s, %s, %s, %s)
                        """,
                        (response.user.id, response.user.email, user_full_name, "member")
                    )
                    conn.commit()
                    cursor.close()
                    user_role = "member"
        except Exception as e:
            print(f"⚠️ Database error: {e}")
            print(f"Using default role: member")
            user_role = "member"
        
        print(f"Final role: {user_role}")
        print(f"========================")
        
        return AuthResponse(
            access_token=response.session.access_token,
            token_type="bearer",
            user={
                "id": response.user.id,
                "email": response.user.email,
                "full_name": user_full_name,
                "role": user_role,
                "user_metadata": {"role": user_role, "full_name": user_full_name}
            },
            message="Sign in successful"
        )

    except HTTPException:
        raise
    except Exception as e:
        error_msg = str(e).lower()
        print(f"Sign in error: {error_msg}")  # Debug logging
        if DEBUG_AUTH:
            print(f"[AUTH DEBUG] Sign-in exception: {e}")
        
        # Check for common Supabase errors
        if "email not confirmed" in error_msg or "email_not_confirmed" in error_msg:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Email not verified. Please check your inbox and click the verification link before signing in."
            )
        elif "invalid login credentials" in error_msg or "invalid credentials" in error_msg:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid email or password"
            )
        elif "user not found" in error_msg:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="No account found with this email"
            )
        
        # Generic error with more details for debugging
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail=f"Authentication failed: {str(e)}"
        )


@app.post("/auth/signout")
async def sign_out():
    """
    Sign out the current user
    """
    try:
        supabase.auth.sign_out()
        return {"message": "Sign out successful"}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Sign out failed: {str(e)}"
        )


@app.get("/api/profile")
async def get_profile(user = Depends(get_current_user)):
    """
    Get current user profile (requires authentication)
    """
    # Get role from database profiles table for consistency
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            cursor.execute(
                "SELECT role, full_name FROM profiles WHERE id = %s",
                (user.user.id,)
            )
            profile = cursor.fetchone()
            cursor.close()
            
            if profile:
                user_role = profile.get("role", "member")
                user_full_name = profile.get("full_name") or user.user.user_metadata.get("full_name", "")
            else:
                # Fallback to metadata if profile not found
                user_role = user.user.user_metadata.get("role", "member")
                user_full_name = user.user.user_metadata.get("full_name", "")
    except Exception as e:
        print(f"⚠️ Profile fetch error: {e}")
        user_role = user.user.user_metadata.get("role", "member")
        user_full_name = user.user.user_metadata.get("full_name", "")
    
    return {
        "id": user.user.id,
        "email": user.user.email,
        "role": user_role,
        "full_name": user_full_name,
        "created_at": user.user.created_at
    }


def auto_expire_sessions():
    """Auto-expire attendance sessions that haven't checked out within 2 hours (120 minutes)"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                UPDATE attendance
                SET status = 'absent', check_out_time = check_in_time + INTERVAL '120 minutes'
                WHERE check_out_time IS NULL
                AND check_in_time < NOW() - INTERVAL '120 minutes'
                AND status != 'absent'
                """
            )
            expired_count = cursor.rowcount
            conn.commit()
            cursor.close()
            return expired_count
    except Exception as e:
        print(f"[ERROR] Failed to auto-expire sessions: {str(e)}")
        return 0


@app.get("/api/admin/dashboard")
async def admin_dashboard(user = Depends(require_admin)):
    """
    Admin-only dashboard endpoint
    Auto-expires sessions that haven't checked out within 2 hours
    """
    try:
        # Auto-expire old sessions before fetching data
        auto_expire_sessions()
        
        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)

            cursor.execute("SELECT COUNT(*) AS total_users, SUM(CASE WHEN role = 'admin' THEN 1 ELSE 0 END) AS total_admins, SUM(CASE WHEN role = 'member' THEN 1 ELSE 0 END) AS total_members FROM profiles")
            profile_counts = cursor.fetchone() or {"total_users": 0, "total_admins": 0, "total_members": 0}

            cursor.execute("SELECT COUNT(*) AS today_attendance FROM attendance WHERE check_in_time >= DATE_TRUNC('day', NOW())")
            today_attendance = cursor.fetchone() or {"today_attendance": 0}

            cursor.execute("SELECT id, email, full_name, role, created_at FROM profiles ORDER BY created_at DESC LIMIT 10")
            recent_users = cursor.fetchall() or []

            cursor.execute(
                """
                SELECT a.id, a.user_id, a.check_in_time, a.check_out_time, a.status, a.notes, a.latitude, a.longitude, a.created_at,
                       p.full_name, p.email
                FROM attendance a
                JOIN profiles p ON a.user_id = p.id
                ORDER BY a.check_in_time DESC
                LIMIT 10
                """
            )
            recent_attendance = cursor.fetchall() or []

            cursor.close()

        return {
            "message": "Welcome to admin dashboard",
            "user_role": user.user.user_metadata.get("role"),
            "stats": {
                "total_users": profile_counts.get("total_users", 0),
                "total_admins": profile_counts.get("total_admins", 0),
                "total_members": profile_counts.get("total_members", 0),
                "today_attendance": today_attendance.get("today_attendance", 0)
            },
            "users": recent_users,
            "recent_attendance": recent_attendance
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to load admin dashboard: {str(e)}"
        )


@app.get("/api/admin/users")
async def get_all_users(user = Depends(require_admin)):
    """
    Get all users/employees (Admin only)
    """
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            cursor.execute(
                """
                SELECT id, email, full_name, role, created_at, updated_at
                FROM profiles
                ORDER BY created_at DESC
                """
            )
            all_users = cursor.fetchall() or []
            cursor.close()

        return {
            "message": "All employees list",
            "total": len(all_users),
            "users": all_users
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to load users: {str(e)}"
        )


def auto_expire_sessions():
    """Auto-expire attendance sessions that haven't checked out within 2 hours (120 minutes)"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                UPDATE attendance
                SET status = 'absent', check_out_time = check_in_time + INTERVAL '120 minutes'
                WHERE check_out_time IS NULL
                AND check_in_time < NOW() - INTERVAL '120 minutes'
                AND status != 'absent'
                """
            )
            expired_count = cursor.rowcount
            conn.commit()
            cursor.close()
            return expired_count
    except Exception as e:
        print(f"[ERROR] Failed to auto-expire sessions: {str(e)}")
        return 0


@app.get("/api/member/attendance")
async def member_attendance(user = Depends(get_current_user)):
    """
    Get member's attendance records (requires authentication)
    Auto-expires sessions that haven't checked out within 2 hours
    """
    try:
        # Auto-expire old sessions before fetching data
        auto_expire_sessions()
        
        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            cursor.execute(
                """
                SELECT id, user_id, check_in_time, check_out_time, status, notes, latitude, longitude, created_at
                FROM attendance
                WHERE user_id = %s
                ORDER BY check_in_time DESC
                """,
                (user.user.id,)
            )
            attendance = cursor.fetchall() or []
            cursor.close()

        stats = {
            "totalDays": len(attendance),
            "presentDays": len([r for r in attendance if r["status"] == "present"]),
            "lateDays": len([r for r in attendance if r["status"] == "late"]),
            "absentDays": len([r for r in attendance if r["status"] == "absent"]),
        }

        return {
            "message": "Member attendance records",
            "user_id": user.user.id,
            "role": user.user.user_metadata.get("role", "member"),
            "attendance": attendance,
            "stats": stats,
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to load attendance: {str(e)}"
        )


@app.post("/api/member/attendance/checkin")
async def check_in(location: CheckInRequest, request: Request, user = Depends(get_current_user), office_ip: str = Depends(require_office_network)):
    """Create a check-in record for the current user (requires office network access)"""
    try:
        # Auto-expire old sessions before creating new check-in
        auto_expire_sessions()
        
        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)

            cursor.execute(
                """
                SELECT id FROM attendance
                WHERE user_id = %s AND check_in_time >= DATE_TRUNC('day', NOW())
                ORDER BY check_in_time DESC
                LIMIT 1
                """,
                (user.user.id,),
            )
            existing = cursor.fetchone()
            if existing:
                cursor.close()
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Already checked in today"
                )

            cursor.execute("SELECT NOW() > DATE_TRUNC('day', NOW()) + INTERVAL '9 hour' AS is_late")
            is_late = cursor.fetchone().get("is_late", False)
            status_value = "late" if is_late else "present"

            cursor.execute(
                """
                INSERT INTO attendance (user_id, status, check_in_time, latitude, longitude)
                VALUES (%s, %s, NOW(), %s, %s)
                RETURNING id, user_id, check_in_time, check_out_time, status, notes, latitude, longitude, created_at
                """,
                (user.user.id, status_value, location.latitude, location.longitude),
            )
            record = cursor.fetchone()
            conn.commit()
            cursor.close()

        return {"message": "Check-in successful", "record": record}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to check in: {str(e)}"
        )


@app.post("/api/member/attendance/checkout")
async def check_out(payload: AttendanceCheckoutRequest, user = Depends(get_current_user)):
    """Update checkout time for a given attendance record"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)

            cursor.execute(
                """
                UPDATE attendance
                SET check_out_time = NOW()
                WHERE id = %s AND user_id = %s AND check_out_time IS NULL
                RETURNING id, user_id, check_in_time, check_out_time, status, notes, latitude, longitude, created_at
                """,
                (payload.attendance_id, user.user.id),
            )
            record = cursor.fetchone()
            if not record:
                cursor.close()
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="No open check-in found for this record"
                )
            conn.commit()
            cursor.close()

        return {"message": "Check-out successful", "record": record}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to check out: {str(e)}"
        )


# Leave Management Endpoints

@app.post("/api/member/leaves")
async def request_leave(payload: LeaveRequest, user = Depends(get_current_user)):
    """Member requests a leave"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            cursor.execute(
                """
                INSERT INTO leaves (user_id, start_date, end_date, reason, status)
                VALUES (%s, %s, %s, %s, 'pending')
                RETURNING id, user_id, start_date, end_date, reason, status, admin_note, approved_by, created_at, updated_at
                """,
                (user.user.id, payload.start_date, payload.end_date, payload.reason),
            )
            leave = cursor.fetchone()
            conn.commit()
            cursor.close()

        return {"message": "Leave request submitted successfully", "leave": leave}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to request leave: {str(e)}"
        )


@app.get("/api/member/leaves")
async def get_member_leaves(user = Depends(get_current_user)):
    """Get all leave requests for the current member"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            cursor.execute(
                """
                SELECT l.id, l.user_id, l.start_date, l.end_date, l.reason, l.status, 
                       l.admin_note, l.approved_by, l.created_at, l.updated_at,
                       p.full_name as approved_by_name
                FROM leaves l
                LEFT JOIN profiles p ON l.approved_by = p.id
                WHERE l.user_id = %s
                ORDER BY l.created_at DESC
                """,
                (user.user.id,),
            )
            leaves = cursor.fetchall() or []
            cursor.close()

        return {"leaves": leaves}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch leaves: {str(e)}"
        )


@app.get("/api/admin/leaves")
async def get_all_leaves(user = Depends(require_admin)):
    """Admin gets all leave requests"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            cursor.execute(
                """
                SELECT l.id, l.user_id, l.start_date, l.end_date, l.reason, l.status, 
                       l.admin_note, l.approved_by, l.created_at, l.updated_at,
                       p.full_name as user_name, p.email as user_email,
                       ap.full_name as approved_by_name
                FROM leaves l
                JOIN profiles p ON l.user_id = p.id
                LEFT JOIN profiles ap ON l.approved_by = ap.id
                ORDER BY 
                    CASE l.status 
                        WHEN 'pending' THEN 1 
                        WHEN 'approved' THEN 2 
                        WHEN 'rejected' THEN 3 
                    END,
                    l.created_at DESC
                """
            )
            leaves = cursor.fetchall() or []
            cursor.close()

        return {"leaves": leaves}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch leaves: {str(e)}"
        )


@app.post("/api/admin/leaves/approve")
async def approve_leave(payload: LeaveApprovalRequest, user = Depends(require_admin)):
    """Admin approves or rejects a leave request"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            cursor.execute(
                """
                UPDATE leaves
                SET status = %s, admin_note = %s, approved_by = %s, updated_at = NOW()
                WHERE id = %s
                RETURNING id, user_id, start_date, end_date, reason, status, admin_note, approved_by, created_at, updated_at
                """,
                (payload.status, payload.admin_note, user.user.id, payload.leave_id),
            )
            leave = cursor.fetchone()
            
            if not leave:
                cursor.close()
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Leave request not found"
                )
            
            conn.commit()
            cursor.close()

        return {"message": f"Leave request {payload.status}", "leave": leave}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update leave: {str(e)}"
        )


# Excel Export Endpoints (Admin Only)

@app.get("/api/admin/export/attendance")
async def export_attendance(user = Depends(require_admin)):
    """Export all attendance records to Excel (Admin only)"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            cursor.execute(
                """
                SELECT a.id, a.user_id, a.check_in_time, a.check_out_time, a.status, a.notes, a.latitude, a.longitude, a.created_at,
                       p.full_name, p.email
                FROM attendance a
                JOIN profiles p ON a.user_id = p.id
                ORDER BY a.check_in_time DESC
                """
            )
            records = cursor.fetchall() or []
            cursor.close()

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        # Define styles
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = ["Employee Name", "Email", "Check-In Date", "Check-In Time", "Check-Out Time", "Status", "Duration (hrs)", "Latitude", "Longitude", "Notes"]
        ws.append(headers)
        
        # Style header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Add data
        for record in records:
            check_in = datetime.fromisoformat(str(record['check_in_time']))
            check_out_time = ""
            duration = ""
            
            if record['check_out_time']:
                check_out = datetime.fromisoformat(str(record['check_out_time']))
                check_out_time = check_out.strftime("%H:%M:%S")
                duration = round((check_out - check_in).total_seconds() / 3600, 2)
            
            ws.append([
                record['full_name'] or 'N/A',
                record['email'],
                check_in.strftime("%Y-%m-%d"),
                check_in.strftime("%H:%M:%S"),
                check_out_time,
                record['status'].capitalize(),
                duration,
                record.get('latitude', 'N/A'),
                record.get('longitude', 'N/A'),
                record['notes'] or ""
            ])

        # Style data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 30

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        filename = f"attendance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to export attendance: {str(e)}"
        )


@app.get("/api/admin/export/leaves")
async def export_leaves(user = Depends(require_admin)):
    """Export all leave requests to Excel (Admin only)"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            cursor.execute(
                """
                SELECT l.id, l.user_id, l.start_date, l.end_date, l.reason, l.status, 
                       l.admin_note, l.approved_by, l.created_at, l.updated_at,
                       p.full_name as user_name, p.email as user_email,
                       ap.full_name as approved_by_name
                FROM leaves l
                JOIN profiles p ON l.user_id = p.id
                LEFT JOIN profiles ap ON l.approved_by = ap.id
                ORDER BY l.created_at DESC
                """
            )
            records = cursor.fetchall() or []
            cursor.close()

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Leave Requests"

        # Define styles
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        # Status colors
        pending_fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
        approved_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        rejected_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = ["Employee Name", "Email", "Start Date", "End Date", "Days", "Reason", "Status", "Admin Note", "Approved By", "Request Date"]
        ws.append(headers)
        
        # Style header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Add data
        for record in records:
            start_date = datetime.fromisoformat(str(record['start_date']))
            end_date = datetime.fromisoformat(str(record['end_date']))
            days = (end_date - start_date).days + 1
            created_at = datetime.fromisoformat(str(record['created_at']))
            
            ws.append([
                record['user_name'] or 'N/A',
                record['user_email'],
                start_date.strftime("%Y-%m-%d"),
                end_date.strftime("%Y-%m-%d"),
                days,
                record['reason'],
                record['status'].capitalize(),
                record['admin_note'] or "",
                record['approved_by_name'] or "",
                created_at.strftime("%Y-%m-%d %H:%M")
            ])

        # Style data rows with status colors
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            status_val = ws.cell(row_idx, 7).value.lower() if ws.cell(row_idx, 7).value else ""
            
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Apply status-based background color
                if status_val == "pending":
                    cell.fill = pending_fill
                elif status_val == "approved":
                    cell.fill = approved_fill
                elif status_val == "rejected":
                    cell.fill = rejected_fill

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 30
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 18

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        filename = f"leave_requests_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to export leaves: {str(e)}"
        )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
